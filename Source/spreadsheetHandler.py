import os
import sys
from os.path import dirname
from openpyxl import load_workbook

# method to verify workbooks. returns list of wb objects
def verifyWB():
    # path name + expected workbooks
    wbPath = dirname(dirname(__file__)) + '/InputFolder/'
    queryWBTitle = 'Medrio_QueryExport_Grail_CCGA.xlsx'
    subjectDataWBTitle = 'Subject_Data_Summary.xlsx'
    invoiceWBTitle = 'CCGAInvoicesTracker.xlsx'
    cancerStatusWBTitle = 'l_enrolled_subj.xlsx'
    mainInputWBTitle = 'inputFile.xlsx'

    wbNames = [queryWBTitle, subjectDataWBTitle, invoiceWBTitle, cancerStatusWBTitle, mainInputWBTitle]
    
    missingWB = False

    # verify workbooks exist
    for wb in wbNames:
        if not os.path.exists(wbPath + wb):
            missingWB = True
            print('Missing Excel workbook:', wb)

    # if workbook doesn't exist, quit program
    if missingWB:
        print('Missing Excel Workbook(s).  Quitting Program')
        sys.exit()
    
    # load workbooks and return as list
    else:
        queryWB = load_workbook(wbPath + wbNames[0])
        subjectDataWB = load_workbook(wbPath + wbNames[1])
        invoicesWB = load_workbook(wbPath + wbNames[2])
        cancerStatusWB = load_workbook(wbPath + wbNames[3])
        mainWB = load_workbook(wbPath + wbNames[4])

        workbooks = [queryWB, subjectDataWB, invoicesWB, cancerStatusWB, mainWB]
        return workbooks


# verifying expected worksheets exist in workbooks
def verifyWorksheets(workbookList):

    # expected worksheets
    queryWSTitle = 'Queries'
    subjectDataWSTitle = 'Subject_Data_Summary.csv'
    invoicesWSTitle = 'Patient Payment'
    cancerStatusWSTitle = 'l_enrolled_subj'
    mainWSTitle = 'Sheet1'

    worksheetList = [queryWSTitle, subjectDataWSTitle, invoicesWSTitle, cancerStatusWSTitle, mainWSTitle]

    missingWS = False

    # check workbooksheets and load
    worksheets = []
    for idx in range(len(workbookList)):
        if not worksheetList[idx] in workbookList[idx].sheetnames:
            print('Missing worksheet - ' + str(worksheetList[idx]) + ' in Workbook - ' + str(workbookList[idx]))
            missingWS = True
        else:
            tempWorkbook = workbookList[idx]
            worksheets.append(tempWorkbook[worksheetList[idx]])
    
    if missingWS:
        print('Missing worksheet(s). Exiting Program.')
        sys.exit()

    return worksheets

def cleanFiles():
    if os.path.exists('./errorLog.txt'):
        os.remove('./errorLog.txt')

def createWorksheet(workbook, worksheetName):
  subjectDataSheet = workbook.create_sheet(title=worksheetName)
  return subjectDataSheet
